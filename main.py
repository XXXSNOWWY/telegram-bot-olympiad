import telebot
from telebot import types
import openpyxl
import os
import re

# Token (Railway environment variable)
TOKEN = os.getenv("TOKEN")
bot = telebot.TeleBot(TOKEN)

# Kanal username
CHANNEL_USERNAME = "@Xamidjonov_Xusniddin"
# Global o'zgaruvchi
waiting_for_broadcast = False

# Faqat admin /sendall buyrug'i
@bot.message_handler(commands=['sendall'])
def send_all_command(message):
    global waiting_for_broadcast
    if message.from_user.id == ADMIN_ID:
        bot.send_message(message.chat.id, "✍️ Barcha foydalanuvchilarga yuboriladigan xabarni kiriting:")
        waiting_for_broadcast = True
    else:
        bot.send_message(message.chat.id, "❌ Sizda ruxsat yo‘q.")

# Admin yuborgan xabarni barchaga tarqatish
@bot.message_handler(func=lambda msg: True)
def handle_message(message):
    global waiting_for_broadcast
    if message.from_user.id == ADMIN_ID and waiting_for_broadcast:
        broadcast_message = message.text

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active

        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Bizda telefon saqlanadi, lekin Telegram ID yo'q
            # Agar Telegram ID saqlasangiz, shu yerda ishlatamiz
            try:
                bot.send_message(row[4], broadcast_message)  # masalan 5-ustunda user_id saqlansa
                count += 1
            except Exception as e:
                print("Xabar yuborilmadi:", e)

        bot.send_message(message.chat.id, f"✅ Xabar {count} ta foydalanuvchiga yuborildi.")
        waiting_for_broadcast = False

# Excel fayl nomi
EXCEL_FILE = "registratsiya.xlsx"

# Excel fayl yo‘q bo‘lsa, yaratamiz
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Qatnashuvchilar"
    sheet.append(["Ism", "Familiya", "Sinf", "Telefon"])
    wb.save(EXCEL_FILE)

# Admin ID
ADMIN_ID = 1302280468

# Foydalanuvchilar ma’lumotlari
user_data = {}

# Obuna tekshirish
def is_subscribed(user_id):
    try:
        member = bot.get_chat_member(CHANNEL_USERNAME, user_id)
        return member.status in ["member", "administrator", "creator"]
    except Exception as e:
        print("is_subscribed xatosi:", e)
        return False


# START komandasi
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id

    # Agar admin bo‘lsa → ro‘yxatni chiqaramiz
    if user_id == ADMIN_ID:
        send_registered_users(message)
        return

    user_data[user_id] = {}

    if is_subscribed(user_id):
        bot.send_message(user_id, "Assalomu alaykum! Ro'yxatdan o‘tishni boshlaymiz.\nIsmingizni yozing:")
        bot.register_next_step_handler(message, get_name)
    else:
        markup = types.InlineKeyboardMarkup()
        join_button = types.InlineKeyboardButton("📢 Kanalga qo‘shilish", url=f"https://t.me/{CHANNEL_USERNAME[1:]}") 
        check_button = types.InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_sub")
        markup.add(join_button)
        markup.add(check_button)
        bot.send_message(user_id, "Avval kanalga obuna bo‘ling 👇", reply_markup=markup)


# Obunani tekshirish tugmasi
@bot.callback_query_handler(func=lambda call: call.data == "check_sub")
def check_subscription(call):
    user_id = call.from_user.id
    if is_subscribed(user_id):
        bot.edit_message_text("✅ Obuna tasdiqlandi! Endi ro‘yxatdan o‘tamiz.\nIsmingizni yozing:",
                              chat_id=call.message.chat.id,
                              message_id=call.message.message_id)
        bot.register_next_step_handler(call.message, get_name)
    else:
        bot.answer_callback_query(call.id, "❌ Siz hali obuna bo‘lmadingiz!")


# Ism
def get_name(message):
    chat_id = message.from_user.id
    text = message.text.strip()

    if not re.match(r'^[A-Z][a-zA-Z\'\- ]+$', text):
        bot.send_message(chat_id, "❌ Ism faqat lotin harflarida va katta harf bilan boshlanishi kerak. Qaytadan kiriting:")
        bot.register_next_step_handler(message, get_name)
        return

    user_data[chat_id]["ism"] = text
    bot.send_message(chat_id, "Familiyangizni yozing:")
    bot.register_next_step_handler(message, get_surname)


# Familiya
def get_surname(message):
    chat_id = message.from_user.id
    text = message.text.strip()

    if not re.match(r'^[A-Z][a-zA-Z\'\- ]+$', text):
        bot.send_message(chat_id, "❌ Familiya faqat lotin harflarida va katta harf bilan boshlanishi kerak. Qaytadan kiriting:")
        bot.register_next_step_handler(message, get_surname)
        return

    user_data[chat_id]["familiya"] = text

    # faqat raqamli klaviatura chiqishi uchun
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("7", "8", "9")
    markup.row("10", "11")
    bot.send_message(chat_id, "Nechanchi sinfda o‘qiysiz?", reply_markup=markup)
    bot.register_next_step_handler(message, get_class)


# Sinf
def get_class(message):
    chat_id = message.from_user.id
    text = message.text.strip()

    if not text.isdigit():
        bot.send_message(chat_id, "❌ Sinf raqamini faqat sonlarda kiriting (masalan: 7 yoki 11).")
        bot.register_next_step_handler(message, get_class)
        return

    user_data[chat_id]["sinf"] = text

    # telefon raqam uchun tugma
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    btn1 = types.KeyboardButton("📱 Raqamni yuborish", request_contact=True)
    markup.add(btn1)
    bot.send_message(chat_id, "Telefon raqamingizni yuboring:", reply_markup=markup)


# Telefon (kontakt orqali)
@bot.message_handler(content_types=['contact'])
def get_contact(message):
    chat_id = message.from_user.id
    phone = message.contact.phone_number.strip()

    # Faqat raqamlarni qoldiramiz
    phone_clean = re.sub(r'\D', '', phone)

    if not re.match(r'^\d{9,15}$', phone_clean):
        bot.send_message(chat_id, "❌ Telefon raqamni to‘g‘ri yuboring. Masalan: +998901234567")
        return

    user_data[chat_id]["telefon"] = "+" + phone_clean  # + qo‘shib qo‘yamiz

    save_to_excel(chat_id)
    bot.send_message(chat_id, "✅ Siz muvaffaqiyatli ro'yxatdan o'tdingiz!", reply_markup=types.ReplyKeyboardRemove())
    
# Excelga yozish
def save_to_excel(chat_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    sheet.append([
        user_data[chat_id]["ism"],
        user_data[chat_id]["familiya"],
        user_data[chat_id]["sinf"],
        user_data[chat_id]["telefon"],
        chat_id   # Telegram foydalanuvchi ID ham saqlanadi
    ])
    wb.save(EXCEL_FILE)

# Admin uchun ro‘yxatni matn ko‘rinishida yuborish
def send_registered_users(message):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        users.append(f"{row[0]} {row[1]}, {row[2]}-sinf, 📱 {row[3]}")

    if users:
        text = "\n".join(users)
        bot.send_message(message.chat.id, f"📋 Ro'yxatdan o'tganlar:\n\n{text}")
    else:
        bot.send_message(message.chat.id, "❌ Hali hech kim ro'yxatdan o'tmagan.")

    # Excel faylni ham yuboradi
    with open(EXCEL_FILE, "rb") as f:
        bot.send_document(message.chat.id, f)


# Faqat admin Excel faylni olish
@bot.message_handler(commands=['getfile'])
def send_file(message):
    if message.from_user.id == ADMIN_ID:
        with open(EXCEL_FILE, "rb") as f:
            bot.send_document(message.chat.id, f)
    else:
        bot.send_message(message.chat.id, "❌ Sizda ruxsat yo‘q.")


print("Bot ishlayapti...")
bot.infinity_polling()
